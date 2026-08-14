# -*- coding: utf-8 -*-
"""
WeaponData.xlsx -> Weapon.json / WeaponTypeData.json / Rarity.json / Localization_{lang}.json

시트 구성:
  - "WeaponTypes"    : 무기 타입 메타데이터 마스터
  - "Rarity"         : 등급 메타데이터 마스터 (SellValue = 판매가의 유일한 출처)
  - "Localization"   : NameKey/DescKey가 참조하는 다국어 텍스트 마스터
  - 그 외 시트(위 3개 + Legend 제외)  : 개별 무기 데이터. 시트 이름 = WeaponType 값과 일치해야 함.

검증 항목:
  1) 헤더 검증     : 공통 필드가 정의된 순서/이름대로 존재하는지
  2) ID 중복 검사   : Index가 전체 시트를 통틀어 유일한지
  3) 참조 검사     : WeaponType / Rarity / NameKey / DescKey가 각각의 마스터 시트에 실제로 존재하는지
  4) 번역 누락 검사 : Localization 시트의 모든 키가 등록된 모든 언어 컬럼에 값을 갖고 있는지
"""

import json
import sys
from pathlib import Path

import openpyxl

# ============================================================
# 단일 출처 스키마 정의 (Single Source of Truth)
# ============================================================
COMMON_FIELDS = [
    "Index", "WeaponType", "Rarity", "NameKey", "DescKey",
    "Damage", "PelletCount", "HeadshotMultiplier", "FireRate_RPS", "FireMode", "BurstCount", "BurstShotInterval",
    "MagazineSize", "ReloadTime",
    "MaxRange", "DamageFalloffStart", "DamageFalloffEnd", "DamageFalloffMin",
    "IsHitscan", "ProjectileSpeed", "CanADS", "ScopeZoomLevel",
    "SpreadHipfire", "SpreadADS", "SpreadIncreasePerShot",
    "MaxSpreadBloomHipfire", "MaxSpreadBloomADS", "SpreadRecoveryDelay", "SpreadRecoveryRate",
    "RecoilVertical", "RecoilHorizontal", "ADSSpeed", "ADSMoveSpeedMultiplier",
    "MoveSpeedMultiplier", "EquipTime",
]

TYPE_MASTER_SHEET = "WeaponTypes"
RARITY_MASTER_SHEET = "Rarity"
LOC_MASTER_SHEET = "Localization"
IGNORE_SHEETS = {"Legend"}
DEFAULTS_BY_TYPE = {"int": 0, "float": 0.0, "bool": False, "string": "", "enum": ""}


class ConversionError(Exception):
    pass


def cast_value(raw, dtype):
    if raw is None:
        return DEFAULTS_BY_TYPE.get(dtype)
    if dtype == "bool":
        return raw if isinstance(raw, bool) else str(raw).strip().upper() == "TRUE"
    if dtype == "int":
        try:
            return int(raw)
        except (ValueError, TypeError):
            return 0
    if dtype == "float":
        try:
            return float(raw)
        except (ValueError, TypeError):
            return 0.0
    return str(raw).strip()


def read_master_sheet(wb, sheet_name, required_fields, key_field):
    if sheet_name not in wb.sheetnames:
        raise ConversionError(f"'{sheet_name}' 시트를 찾을 수 없습니다.")
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    dtypes = [c.value for c in ws[2]]
    col_map = {h: i for i, h in enumerate(headers)}
    missing = [f for f in required_fields if f not in col_map]
    if missing:
        raise ConversionError(f"{sheet_name} 헤더 누락: {missing}")

    master, ordered = {}, []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[col_map[key_field]] is None:
            continue
        entry = {f: cast_value(row[col_map[f]], dtypes[col_map[f]]) for f in required_fields}
        key = entry[key_field]
        if key in master:
            raise ConversionError(f"{sheet_name} 내 {key_field} 중복: '{key}'")
        master[key] = entry
        ordered.append(entry)
    return master, ordered


def read_localization_sheet(wb):
    if LOC_MASTER_SHEET not in wb.sheetnames:
        raise ConversionError(f"'{LOC_MASTER_SHEET}' 시트를 찾을 수 없습니다.")
    ws = wb[LOC_MASTER_SHEET]
    headers = [c.value for c in ws[1]]
    if not headers or headers[0] != "Key":
        raise ConversionError("Localization 시트의 첫 컬럼은 'Key'여야 합니다.")

    language_columns = [h for h in headers[1:] if h is not None]
    if not language_columns:
        raise ConversionError("Localization 시트에 언어 컬럼이 하나도 없습니다.")

    loc_master = {}
    errors = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[0]
        if key is None:
            continue
        if key in loc_master:
            errors.append(f"Localization 내 Key 중복: '{key}'")
            continue
        translations = {}
        for i, lang in enumerate(language_columns, start=1):
            value = row[i] if i < len(row) else None
            if value is None or str(value).strip() == "":
                errors.append(f"[Localization/{key}] '{lang}' 번역이 비어있습니다.")
            translations[lang] = str(value).strip() if value is not None else ""
        loc_master[key] = translations

    return loc_master, language_columns, errors


def discover_weapon_sheets(wb):
    reserved = {TYPE_MASTER_SHEET, RARITY_MASTER_SHEET, LOC_MASTER_SHEET} | IGNORE_SHEETS
    return [n for n in wb.sheetnames if n not in reserved]


def read_weapon_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    dtypes = [c.value for c in ws[2]]
    prefix = headers[:len(COMMON_FIELDS)]
    if prefix != COMMON_FIELDS:
        raise ConversionError(f"[{sheet_name}] 헤더 검증 실패.\n  기대값: {COMMON_FIELDS}\n  실제값: {prefix}")
    col_dtype = {headers[i]: dtypes[i] for i in range(len(headers)) if headers[i] is not None}
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        entry = {headers[i]: cast_value(row[i], col_dtype[headers[i]])
                 for i in range(len(headers)) if headers[i] is not None}
        rows.append(entry)
    return rows, col_dtype


def convert(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    type_master, type_list = read_master_sheet(
        wb, TYPE_MASTER_SHEET, ["TypeID", "Category", "DisplayName", "Description", "IconPath"], "TypeID")
    rarity_master, rarity_list = read_master_sheet(
        wb, RARITY_MASTER_SHEET, ["RarityID", "DisplayName", "Color", "DropWeight", "SellValue"], "RarityID")
    loc_master, language_columns, loc_errors = read_localization_sheet(wb)

    weapon_sheets = discover_weapon_sheets(wb)
    if not weapon_sheets:
        raise ConversionError("무기 데이터 시트를 하나도 찾지 못했습니다.")

    all_rows, global_dtype_map, errors, seen_index = [], {}, list(loc_errors), {}

    for sheet_name in weapon_sheets:
        try:
            rows, col_dtype = read_weapon_sheet(wb, sheet_name)
        except ConversionError as e:
            errors.append(str(e))
            continue

        global_dtype_map.update(col_dtype)

        if sheet_name not in type_master:
            errors.append(f"[{sheet_name}] WeaponTypes 마스터에 TypeID='{sheet_name}'가 없습니다.")
            continue

        for entry in rows:
            idx = entry.get("Index")
            wtype = entry.get("WeaponType")
            rarity = entry.get("Rarity")
            name_key = entry.get("NameKey")
            desc_key = entry.get("DescKey")

            if idx in seen_index:
                errors.append(f"Index 중복: '{idx}' ({seen_index[idx]} / {sheet_name})")
            else:
                seen_index[idx] = sheet_name

            if wtype not in type_master:
                errors.append(f"[{sheet_name}/{idx}] WeaponType '{wtype}' 미등록")
            elif wtype != sheet_name:
                errors.append(f"[{sheet_name}/{idx}] WeaponType('{wtype}') != 시트명('{sheet_name}')")

            if rarity not in rarity_master:
                errors.append(f"[{sheet_name}/{idx}] Rarity '{rarity}' 이 Rarity 마스터에 없습니다.")

            if name_key not in loc_master:
                errors.append(f"[{sheet_name}/{idx}] NameKey '{name_key}' 이 Localization 시트에 없습니다.")
            if desc_key not in loc_master:
                errors.append(f"[{sheet_name}/{idx}] DescKey '{desc_key}' 이 Localization 시트에 없습니다.")

            entry["Category"] = type_master.get(wtype, {}).get("Category", "")
            all_rows.append(entry)

    if errors:
        print("검증 실패:", file=sys.stderr)
        for e in errors:
            print(f"  - {e}", file=sys.stderr)
        raise ConversionError(f"총 {len(errors)}건의 오류로 변환을 중단했습니다.")

    all_field_names, seen_fields = [], set()
    for entry in all_rows:
        for k in entry:
            if k not in seen_fields:
                seen_fields.add(k)
                all_field_names.append(k)

    normalized = []
    for entry in all_rows:
        norm = {}
        for field in all_field_names:
            norm[field] = entry.get(field, DEFAULTS_BY_TYPE.get(global_dtype_map.get(field, "string")))
        normalized.append(norm)

    # 언어별 딕셔너리로 재구성: {"P_1.Name": "Sidearm Mk1", ...}
    localization_by_lang = {lang: {} for lang in language_columns}
    for key, translations in loc_master.items():
        for lang in language_columns:
            localization_by_lang[lang][key] = translations.get(lang, "")

    return type_list, rarity_list, normalized, localization_by_lang


def main():
    xlsx_path = Path(sys.argv[1] if len(sys.argv) > 1 else "WeaponData.xlsx")
    out_dir = Path(sys.argv[2] if len(sys.argv) > 2 else ".")
    out_dir.mkdir(parents=True, exist_ok=True)

    try:
        type_list, rarity_list, weapons, localization_by_lang = convert(xlsx_path)
    except ConversionError as e:
        print(f"\n변환 실패: {e}", file=sys.stderr)
        sys.exit(1)

    (out_dir / "WeaponTypeData.json").write_text(
        json.dumps(type_list, ensure_ascii=False, indent=2), encoding="utf-8")
    (out_dir / "Rarity.json").write_text(
        json.dumps(rarity_list, ensure_ascii=False, indent=2), encoding="utf-8")
    (out_dir / "Weapon.json").write_text(
        json.dumps(weapons, ensure_ascii=False, indent=2), encoding="utf-8")

    for lang, table in localization_by_lang.items():
        (out_dir / f"Localization_{lang}.json").write_text(
            json.dumps(table, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"변환 완료: 타입 {len(type_list)}개, 등급 {len(rarity_list)}개, 무기 {len(weapons)}개, "
          f"언어 {len(localization_by_lang)}개 ({', '.join(localization_by_lang.keys())})")


if __name__ == "__main__":
    main()
