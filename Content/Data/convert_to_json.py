"""
RandomBox.xlsx + 카테고리별 아이템 원본 파일(WeaponData.xlsx 등) -> JSON 컨버터

파일 역할 분리:
  - RandomBox.xlsx  : Rarity(등급 마스터), RandomBox(박스 메타), DropTable(박스별 아이템 가중치),
                       ItemCategory(카테고리 -> 원본 파일 매핑 문서)
  - WeaponData.xlsx 등 카테고리 파일: 그 카테고리 아이템의 실제 스탯/로컬라이징 원본 데이터

DropTable 은 등급별로 분리되어 있다 (DropTable_Common, DropTable_Rare, DropTable_Epic, ...).
각 시트의 행은 (BoxID, ItemCategory, ItemIndex, ItemDropWeight, IsEnabled) 로 구성되며,
시트 이름 자체가 그 행들의 등급을 의미한다. Rarity 는 시트에 저장하지 않는다 -
ItemCategory 로 어느 원본 파일을 봐야 하는지 알고, 그 파일에서 ItemIndex 로 아이템을 찾아
실제 Rarity 를 조회해서 시트 이름과 일치하는지 검증한다 (원본 데이터가 유일한 출처).

추첨은 2단계:
  1단계) Rarity 시트의 DropWeight 로 등급을 선택
  2단계) 그 등급에 속한 아이템들 중 DropTable.ItemDropWeight 로 아이템을 선택

출력물은 카테고리별로 분리된 JSON 이다:
  - rarities.json     : 카테고리 공통 등급 마스터 (단일 출처, 중복 저장 안 함)
  - {category}.json   : 카테고리별 아이템 리스트 (예: weapons.json). 파일명은 CATEGORY_FILE_NAMES 참조.
  - gacha_tables.json : 박스/추첨 전용 경량 데이터

새 카테고리(Skin, Consumable 등)를 추가하려면:
  1. RandomBox.xlsx 의 ItemCategory 시트에 카테고리 한 줄 추가
  2. 그 카테고리 아이템 원본 파일을 CATEGORY_LOADERS 에 등록 (아래 참고)
  3. 출력 파일명을 CATEGORY_FILE_NAMES 에 등록 (안 하면 "{category.lower()}.json"으로 자동 생성)
  4. DropTable 에 해당 카테고리로 행 추가
만 하면 되고, 이 스크립트의 핵심 로직(build_gacha_tables)은 그대로 재사용된다.

사용법:
    python convert_to_json.py RandomBox.xlsx WeaponData.xlsx ./output
    (카테고리 파일은 여러 개 추가 가능: ... RandomBox.xlsx WeaponData.xlsx SkinData.xlsx ./output)
"""
import sys
import json
from pathlib import Path
import openpyxl

WEAPON_SHEETS = ["Pistol", "AssaultRifle", "SniperRifle", "SubMachineGun", "ShotGun"]


def _normalize_value(v):
    """엑셀 드롭다운 검증에 쓴 'TRUE'/'FALSE' 문자열을 실제 JSON bool로 변환.
    (그대로 두면 JSON에 문자열 "TRUE"로 나가서 언리얼의 bool 파싱이 깨짐)"""
    if isinstance(v, str) and v in ("TRUE", "FALSE"):
        return v == "TRUE"
    return v


def sheet_rows(ws, has_type_row=True):
    """1행=헤더, (있다면) 2행=타입, 그 다음부터 데이터. 헤더:값 dict 리스트로 반환."""
    rows = list(ws.iter_rows(values_only=True))
    min_rows = 3 if has_type_row else 2
    if len(rows) < min_rows:
        return []
    headers = rows[0]
    data_start = 2 if has_type_row else 1
    out = []
    for row in rows[data_start:]:
        if row[0] is None:
            continue
        out.append({h: _normalize_value(v) for h, v in zip(headers, row)})
    return out


# ---------------------------------------------------------------------------
# 카테고리별 로더. "이 카테고리 파일을 어떻게 읽어서 {index: item} 으로 만드는가"를 정의.
# 새 카테고리를 추가할 때 이 dict 에 함수 하나만 등록하면 된다.
# ---------------------------------------------------------------------------

def load_weapon_category(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    localization = {}
    for row in sheet_rows(wb["Localization"], has_type_row=False):
        localization[row["Key"]] = {"ko": row["ko"], "en": row["en"]}

    def localize(key):
        entry = localization.get(key)
        return entry if entry is not None else {"ko": key, "en": key}

    items = {}
    for sheet_name in WEAPON_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        for row in sheet_rows(wb[sheet_name]):
            index = row["Index"]
            stats = {k: v for k, v in row.items() if k not in ("Index", "NameKey", "DescKey")}
            items[index] = {
                "index": index,
                "weaponType": row["WeaponType"],
                "rarity": row["Rarity"],
                "name": localize(row["NameKey"]),
                "description": localize(row["DescKey"]),
                "stats": stats,
            }
    return items


# 카테고리 이름 -> 로더 함수. 앞으로 "Skin": load_skin_category 처럼 추가.
CATEGORY_LOADERS = {
    "Weapon": load_weapon_category,
}

# 카테고리 이름 -> 출력 파일명. 카테고리마다 독립된 JSON으로 분리해서 내보낸다.
# (Git 충돌 방지, 언리얼에서 필요한 카테고리만 선택적으로 로드하기 위함)
# 새 카테고리를 추가할 때 CATEGORY_LOADERS와 함께 이 dict에도 등록할 것.
CATEGORY_FILE_NAMES = {
    "Weapon": "weapons.json",
}


def load_randombox_data(randombox_path):
    wb = openpyxl.load_workbook(randombox_path, data_only=True)

    rarities = {}
    for row in sheet_rows(wb["Rarity"]):
        rarities[row["RarityID"]] = {
            "rarityId": row["RarityID"],
            "displayName": row["DisplayName"],
            "color": row["Color"],
            "dropWeight": float(row["DropWeight"]),
            "sellValue": int(row["SellValue"]),
        }

    boxes_meta = {}
    for row in sheet_rows(wb["RandomBox"]):
        boxes_meta[row["BoxID"]] = {
            "displayName": row["DisplayName"],
            "description": row["Description"],
            "isActive": bool(row["IsActive"]),
        }

    # DropTable 은 등급별로 분리되어 있다: DropTable_Common, DropTable_Rare, DropTable_Epic, ...
    # Rarity 시트에 있는 등급마다 해당 이름의 시트를 찾아서 읽는다.
    # 각 행에 그 시트가 의미하는 등급을 sheetRarity 로 붙여서 반환한다 (실제 검증은 build_gacha_tables 에서).
    drop_rows = []
    for rarity_id in rarities:
        sheet_name = f"DropTable_{rarity_id}"
        if sheet_name not in wb.sheetnames:
            continue
        for row in sheet_rows(wb[sheet_name]):
            row["_sheetRarity"] = rarity_id
            drop_rows.append(row)

    return rarities, boxes_meta, drop_rows


def build_gacha_tables(rarities, category_items, boxes_meta, drop_rows):
    """category_items: {"Weapon": {index: item, ...}, "Skin": {...}, ...}"""
    boxes = {}
    warnings = []

    for row in drop_rows:
        box_id = row["BoxID"]
        category = row["ItemCategory"]
        item_index = row["ItemIndex"]
        weight = row["ItemDropWeight"]
        is_enabled = bool(row["IsEnabled"])
        sheet_rarity = row["_sheetRarity"]  # DropTable_<RarityID> 시트 이름이 의미하는 등급

        if not is_enabled:
            continue

        items = category_items.get(category)
        if items is None:
            warnings.append(f"[DropTable_{sheet_rarity}] BoxID={box_id}: ItemCategory '{category}' 에 대한 원본 데이터가 로드되지 않았습니다. 건너뜀.")
            continue

        item = items.get(item_index)
        if item is None:
            warnings.append(f"[DropTable_{sheet_rarity}] BoxID={box_id}: {category}:{item_index} 가 원본 데이터에 없습니다. 건너뜀.")
            continue

        if weight is None or float(weight) <= 0:
            warnings.append(f"[DropTable_{sheet_rarity}] BoxID={box_id}, {category}:{item_index}: ItemDropWeight 가 0 이하입니다. 건너뜀.")
            continue

        # 핵심 검증: 시트 이름(=등급)과 원본 데이터의 실제 Rarity 가 일치해야 한다.
        # 진짜 등급의 출처는 항상 원본 카테고리 파일이므로, 어긋나면 여기서 걸러낸다.
        actual_rarity = item["rarity"]
        if actual_rarity != sheet_rarity:
            warnings.append(
                f"[DropTable_{sheet_rarity}] BoxID={box_id}: {category}:{item_index} 의 실제 Rarity 는 "
                f"'{actual_rarity}' 인데 DropTable_{sheet_rarity} 시트에 들어있습니다. "
                f"올바른 시트(DropTable_{actual_rarity})로 옮겨주세요. 이번 변환에서는 건너뜀."
            )
            continue

        rarity_id = actual_rarity
        if rarity_id not in rarities:
            warnings.append(f"[DropTable_{sheet_rarity}] {category}:{item_index}: Rarity '{rarity_id}' 가 Rarity 시트에 없습니다. 건너뜀.")
            continue

        box = boxes.setdefault(box_id, {
            **boxes_meta.get(box_id, {"displayName": box_id, "description": "", "isActive": True}),
            "pools": {},
        })
        pool = box["pools"].setdefault(rarity_id, {
            "rarityDropWeight": rarities[rarity_id]["dropWeight"],
            "items": [],
        })
        pool["items"].append({"category": category, "index": item_index, "weight": float(weight)})

    for box_id, meta in boxes_meta.items():
        boxes.setdefault(box_id, {**meta, "pools": {}})

    for box_id, box in boxes.items():
        for rarity_id, pool in box["pools"].items():
            if not pool["items"]:
                warnings.append(f"[Box {box_id}] Rarity '{rarity_id}' 풀에 활성 아이템이 없습니다.")

    return {"rarities": rarities, "boxes": boxes}, warnings


def main():
    if len(sys.argv) < 4:
        print("usage: python convert_to_json.py <RandomBox.xlsx> <CategoryFile1.xlsx> [CategoryFile2.xlsx ...] <output_dir>")
        print("  (현재 지원 카테고리 파일: WeaponData.xlsx 형식 -> 'Weapon' 카테고리)")
        sys.exit(1)

    randombox_path = Path(sys.argv[1])
    *category_paths, out_dir = sys.argv[2:]
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    rarities, boxes_meta, drop_rows = load_randombox_data(randombox_path)

    # 현재는 파일 하나 = Weapon 카테고리 하나로 단순 매핑한다.
    # 카테고리가 늘어나면 파일명 규칙이나 별도 인자로 카테고리를 지정하도록 확장하면 된다.
    category_items = {}
    for p in category_paths:
        category_items["Weapon"] = load_weapon_category(p)

    # rarities.json : 카테고리들이 공유하는 등급 마스터 (단일 출처, 카테고리 파일마다 중복 저장하지 않음)
    rarities_path = out_dir / "rarities.json"
    with open(rarities_path, "w", encoding="utf-8") as f:
        json.dump(list(rarities.values()), f, ensure_ascii=False, indent=2)

    # 카테고리별 출력: {카테고리명}.json (예: weapons.json). 리스트(JSON 배열) 형태로 저장 -
    # 언리얼에서 FJsonObjectConverter::JsonArrayStringToUStruct로 그대로 파싱 가능하게 하기 위함.
    category_output_paths = {}
    for category, items in category_items.items():
        filename = CATEGORY_FILE_NAMES.get(category)
        if filename is None:
            # 새 카테고리를 등록할 때 CATEGORY_FILE_NAMES에 파일명을 안 넣은 경우를 위한 안전한 기본값.
            filename = f"{category.lower()}.json"
        path = out_dir / filename
        with open(path, "w", encoding="utf-8") as f:
            json.dump(list(items.values()), f, ensure_ascii=False, indent=2)
        category_output_paths[category] = path

    # gacha_tables.json : 박스/추첨 전용 경량 데이터
    gacha_data, warnings = build_gacha_tables(rarities, category_items, boxes_meta, drop_rows)
    gacha_path = out_dir / "gacha_tables.json"
    with open(gacha_path, "w", encoding="utf-8") as f:
        json.dump(gacha_data, f, ensure_ascii=False, indent=2)

    print(f"[OK] {rarities_path}  (등급 {len(rarities)}개)")
    for category, path in category_output_paths.items():
        print(f"[OK] {path}  ({category} 아이템 {len(category_items[category])}개)")
    print(f"[OK] {gacha_path}  (박스 {len(gacha_data['boxes'])}개)")

    if warnings:
        print("\n[WARN] 확인이 필요한 항목:")
        for w in warnings:
            print("  -", w)
    else:
        print("\n검증 통과: 경고 없음.")


if __name__ == "__main__":
    main()
