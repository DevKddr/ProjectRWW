"""
player_base_stat_converter.py
====================
플레이어 "기본 스탯 설정"(xlsx) 템플릿 생성 + xlsx -> json 변환을 하나의 파일로 처리하는 도구.

설계 배경:
  - 모든 플레이어가 동일한 기본 스탯을 공유하고, 직업/타입 구분이 없다.
  - 향후 퍽/특성 시스템이 붙어도 그건 이 기본 스탯 위에 얹는 별도의 "보정치" 데이터이지,
    기본 스탯 자체가 여러 세트로 늘어나는 게 아니다.
  - 따라서 이 데이터는 "여러 행을 Id로 구분하는 테이블"이 아니라 "단일 설정값(Config)"이다.
    xlsx에는 Id 컬럼이 없고, 정확히 1개의 데이터 행만 존재해야 하며,
    JSON도 배열이 아니라 단일 객체로 출력된다. (언리얼 쪽은 DataTable이 아니라
    런타임에 JSON을 그대로 읽어들이는 단일 구조체/DataAsset 개념으로 대응한다.)

내부적으로는 역할을 함수/클래스 단위로 계속 분리해서 유지한다.
(파일만 하나로 합쳤을 뿐, "하나의 함수가 여러 책임을 갖는" 구조는 아님)

  1. FieldSpec / PLAYER_STAT_SCHEMA   - 스키마 정의 (필드 추가 시 여기만 수정)
  2. build_template()                 - xlsx 템플릿 생성
  3. read_rows()                      - xlsx 읽기 (순수 I/O)
  4. parse_rows()                     - 타입 캐스팅
  5. validate_rows() / validate_single_row_rule() - 검증
  6. export_json()                    - 결과 저장 (단일 객체)
  7. main()                           - CLI 진입점 (template / convert 서브커맨드)

사용법:
    python player_base_stat_converter.py template --output PlayerBaseStat.xlsx
    python player_base_stat_converter.py convert  --input PlayerBaseStat.xlsx --outdir ./output
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from dataclasses import dataclass
from typing import Any, Callable, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# 1. 예외
# ---------------------------------------------------------------------------
class ConverterError(Exception):
    """파이프라인 전반에서 사용하는 베이스 예외."""


class SheetReadError(ConverterError):
    """xlsx 시트를 읽는 도중 발생하는 오류 (파일 없음, 시트 없음, 헤더 형식 불일치 등)."""


# ---------------------------------------------------------------------------
# 2. 스키마 정의
#    Id 없음: 모든 플레이어가 공유하는 기본 스탯이므로 행을 구분할 키가 필요 없다.
#    필드를 추가/변경할 때는 이 리스트와 TEMPLATE_COLUMNS만 함께 수정하면 된다.
# ---------------------------------------------------------------------------
@dataclass
class FieldSpec:
    name: str                              # xlsx 1행(key)과 동일해야 함
    type: type                             # int, float, str
    required: bool = True
    min_value: Optional[float] = None      # 숫자 필드 최소값 (포함)
    max_value: Optional[float] = None      # 숫자 필드 최대값 (포함)
    export: bool = True                    # false면 JSON 변환 시 제외 (예: 기획 메모)
    validator: Optional[Callable[[Any], Optional[str]]] = None


PLAYER_STAT_SCHEMA: list[FieldSpec] = [
    FieldSpec(name="MaxHP", type=float, required=True, min_value=1.0),
    FieldSpec(name="HPRegen", type=float, required=True, min_value=0.0),
    FieldSpec(name="HPRegenDelay", type=float, required=True, min_value=0.0),
    FieldSpec(name="WalkSpeed", type=float, required=True, min_value=0.0),
    FieldSpec(name="RunSpeed", type=float, required=True, min_value=0.0),
    FieldSpec(name="MaxMana", type=float, required=True, min_value=0.0),
    FieldSpec(name="ManaRegen", type=float, required=True, min_value=0.0),
    FieldSpec(name="JumpPower", type=float, required=True, min_value=0.0),
    FieldSpec(name="Comment", type=str, required=False, export=False),
]

# xlsx 템플릿용 표시 정보 (한글 설명, 예시값). PLAYER_STAT_SCHEMA와 순서/이름이 1:1로 맞아야 함.
TEMPLATE_COLUMNS = [
    {"key": "MaxHP",        "header_kr": "최대 체력",                "type": "float",  "example": 100.0},
    {"key": "HPRegen",      "header_kr": "초당 체력 회복량",          "type": "float",  "example": 2.0},
    {"key": "HPRegenDelay", "header_kr": "피격 후 체력회복 지연시간(초)", "type": "float", "example": 5.0},
    {"key": "WalkSpeed",    "header_kr": "걷기 이동속도",             "type": "float",  "example": 300.0},
    {"key": "RunSpeed",     "header_kr": "달리기 이동속도",           "type": "float",  "example": 600.0},
    {"key": "MaxMana",      "header_kr": "최대 마나",                 "type": "float",  "example": 1000.0},
    {"key": "ManaRegen",    "header_kr": "초당 마나 회복량",          "type": "float",  "example": 10.0},
    {"key": "JumpPower",    "header_kr": "점프력",                    "type": "float",  "example": 420.0},
    {"key": "Comment",      "header_kr": "기획 메모(변환시 제외)",   "type": "string", "example": "기본 스탯"},
]

SHEET_NAME = "PlayerBaseStat"
HEADER_ROW = 1        # 실제 파싱 기준이 되는 key 행
TYPE_ROW = 2          # 타입 힌트 행
KR_ROW = 3            # 한글 설명 행
DATA_START_ROW = 4    # 실제 데이터 시작 행


# ---------------------------------------------------------------------------
# 3. 템플릿 생성
# ---------------------------------------------------------------------------
def build_template(path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    type_font = Font(name="Arial", italic=True, color="808080", size=9)
    kr_font = Font(name="Arial", color="404040", size=9)
    editable_fill = PatternFill("solid", fgColor="FFF2CC")

    for col_idx, col in enumerate(TEMPLATE_COLUMNS, start=1):
        letter = get_column_letter(col_idx)

        c = ws.cell(row=HEADER_ROW, column=col_idx, value=col["key"])
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

        t = ws.cell(row=TYPE_ROW, column=col_idx, value=f"[{col['type']}]")
        t.font = type_font
        t.alignment = Alignment(horizontal="center")

        k = ws.cell(row=KR_ROW, column=col_idx, value=col["header_kr"])
        k.font = kr_font
        k.alignment = Alignment(horizontal="center", wrap_text=True)

        e = ws.cell(row=DATA_START_ROW, column=col_idx, value=col["example"])
        e.fill = editable_fill

        ws.column_dimensions[letter].width = max(14, len(col["header_kr"]) + 2)

    ws.freeze_panes = f"A{DATA_START_ROW}"
    ws.row_dimensions[KR_ROW].height = 30

    for key in ("MaxHP", "HPRegen", "HPRegenDelay", "WalkSpeed", "RunSpeed", "MaxMana", "ManaRegen", "JumpPower"):
        col_idx = next(i for i, c in enumerate(TEMPLATE_COLUMNS, start=1) if c["key"] == key)
        letter = get_column_letter(col_idx)
        dv = DataValidation(
            type="decimal", operator="greaterThan", formula1=0,
            error=f"{key}는 0보다 커야 합니다.", errorTitle="범위 오류",
        )
        ws.add_data_validation(dv)
        dv.add(f"{letter}{DATA_START_ROW}:{letter}500")

    guide = wb.create_sheet("README")
    guide["A1"] = "작성 가이드"
    guide["A1"].font = Font(bold=True, size=13)
    lines = [
        "1행: 실제 필드명(key) - 절대 수정/삭제 금지 (변환 스크립트가 이 행을 기준으로 파싱)",
        "2행: 타입 힌트 - 참고용, 수정 불필요",
        "3행: 한글 설명 - 참고용, 수정 불필요",
        "4행: 실제 데이터 (노란색 배경 = 예시 행, 값만 수정)",
        "이 테이블은 모든 플레이어가 공유하는 기본 스탯이므로, 데이터 행은 반드시 1개만 있어야 합니다.",
        "행을 추가/복사하지 마세요. 2행 이상이면 변환 시 오류로 처리됩니다.",
        "Comment 컬럼은 기획 메모용이며 JSON 변환 시 자동으로 제외됩니다.",
    ]
    for i, line in enumerate(lines, start=3):
        guide[f"A{i}"] = f"- {line}"
        guide[f"A{i}"].font = Font(name="Arial", size=10)
    guide.column_dimensions["A"].width = 90

    wb.save(path)
    print(f"템플릿 생성 완료: {path}")


# ---------------------------------------------------------------------------
# 4. xlsx 읽기 (순수 I/O)
# ---------------------------------------------------------------------------
def read_rows(path: str, sheet_name: str, header_row: int, data_start_row: int) -> list[dict]:
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except FileNotFoundError:
        raise SheetReadError(f"파일을 찾을 수 없습니다: {path}")

    if sheet_name not in wb.sheetnames:
        raise SheetReadError(f"시트 '{sheet_name}'를 찾을 수 없습니다. 존재하는 시트: {wb.sheetnames}")

    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[header_row]]
    if not any(headers):
        raise SheetReadError(f"{header_row}행에서 헤더를 찾을 수 없습니다.")

    rows: list[dict] = []
    for row_idx in range(data_start_row, ws.max_row + 1):
        row_values = [cell.value for cell in ws[row_idx]]

        if all(v is None or str(v).strip() == "" for v in row_values):
            continue  # 빈 행은 건너뜀

        row_dict: dict = {}
        for key, value in zip(headers, row_values):
            if key is None:
                continue
            row_dict[str(key)] = value
        row_dict["__excel_row__"] = row_idx
        rows.append(row_dict)

    return rows


# ---------------------------------------------------------------------------
# 5. 타입 캐스팅 (의미 검증은 하지 않음)
# ---------------------------------------------------------------------------
def parse_rows(raw_rows: list[dict], schema: list[FieldSpec]) -> tuple[list[dict], list[str]]:
    parsed_rows: list[dict] = []
    errors: list[str] = []

    for raw in raw_rows:
        excel_row = raw.get("__excel_row__", "?")
        parsed: dict = {}

        for field in schema:
            raw_value = raw.get(field.name)

            if raw_value is None or (isinstance(raw_value, str) and raw_value.strip() == ""):
                parsed[field.name] = None
                continue

            try:
                if field.type is int:
                    parsed[field.name] = int(raw_value)
                elif field.type is float:
                    parsed[field.name] = float(raw_value)
                else:
                    parsed[field.name] = str(raw_value).strip()
            except (ValueError, TypeError):
                errors.append(
                    f"[Excel {excel_row}행] '{field.name}' 값 '{raw_value}' 을(를) "
                    f"{field.type.__name__} 타입으로 변환할 수 없습니다."
                )
                parsed[field.name] = None

        parsed["__excel_row__"] = excel_row
        parsed_rows.append(parsed)

    return parsed_rows, errors


# ---------------------------------------------------------------------------
# 6. 검증
#    6-1) validate_row_count(): 이 테이블 고유 규칙 - 데이터 행이 정확히 1개여야 한다.
#    6-2) validate_rows(): 필드 단위 규칙 (필수값 / 범위 / 커스텀)
# ---------------------------------------------------------------------------
def validate_row_count(parsed_rows: list[dict]) -> list[str]:
    if len(parsed_rows) == 0:
        return ["데이터 행이 없습니다. 4행에 기본 스탯 값을 입력하세요."]
    if len(parsed_rows) > 1:
        excel_rows = [str(r.get("__excel_row__", "?")) for r in parsed_rows[1:]]
        return [
            f"기본 스탯 테이블은 1개의 행만 가져야 합니다. "
            f"{len(parsed_rows)}개의 행이 발견되었습니다 (초과 행: Excel {', '.join(excel_rows)}행). "
            f"불필요한 행을 삭제하세요."
        ]
    return []


def validate_rows(parsed_rows: list[dict], schema: list[FieldSpec]) -> list[str]:
    errors: list[str] = []

    for row in parsed_rows:
        excel_row = row.get("__excel_row__", "?")
        row_label = f"[Excel {excel_row}행]"

        for field in schema:
            value = row.get(field.name)

            if field.required and value is None:
                errors.append(f"{row_label} 필수 필드 '{field.name}'이(가) 비어 있습니다.")
                continue

            if value is None:
                continue

            if field.min_value is not None and isinstance(value, (int, float)) and value < field.min_value:
                errors.append(f"{row_label} '{field.name}' 값 {value}이(가) 최소값 {field.min_value}보다 작습니다.")

            if field.max_value is not None and isinstance(value, (int, float)) and value > field.max_value:
                errors.append(f"{row_label} '{field.name}' 값 {value}이(가) 최대값 {field.max_value}보다 큽니다.")

            if field.validator is not None:
                custom_error = field.validator(value)
                if custom_error:
                    errors.append(f"{row_label} '{field.name}': {custom_error}")

    return errors


# ---------------------------------------------------------------------------
# 7. 결과 저장 - 단일 객체로 출력 (배열 아님, RowName 개념 없음)
# ---------------------------------------------------------------------------
def export_json(single_row: dict, schema: list[FieldSpec], output_path: str) -> None:
    exportable_fields = [f for f in schema if f.export]
    obj = {field.name: single_row.get(field.name) for field in exportable_fields}

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)


# ---------------------------------------------------------------------------
# 8. 오케스트레이션
# ---------------------------------------------------------------------------
def run_convert(input_path: str, outdir: str) -> int:
    try:
        raw_rows = read_rows(input_path, SHEET_NAME, HEADER_ROW, DATA_START_ROW)
    except ConverterError as e:
        print(f"[읽기 실패] {e}")
        return 1

    parsed_rows, type_errors = parse_rows(raw_rows, PLAYER_STAT_SCHEMA)

    # 행 개수 규칙은 필드 검증보다 먼저 확인한다 (0개/2개 이상이면 필드 검증 자체가 무의미하므로)
    row_count_errors = validate_row_count(parsed_rows)
    if row_count_errors:
        print(f"검증 실패: {len(row_count_errors)}건의 문제가 발견되었습니다.\n")
        for e in row_count_errors:
            print(f"  - {e}")
        print("\n변환을 중단합니다. 위 문제를 xlsx에서 수정한 뒤 다시 실행하세요.")
        return 1

    validation_errors = validate_rows(parsed_rows, PLAYER_STAT_SCHEMA)
    all_errors = type_errors + validation_errors
    if all_errors:
        print(f"검증 실패: {len(all_errors)}건의 문제가 발견되었습니다.\n")
        for e in all_errors:
            print(f"  - {e}")
        print("\n변환을 중단합니다. 위 문제를 xlsx에서 수정한 뒤 다시 실행하세요.")
        return 1

    os.makedirs(outdir, exist_ok=True)
    json_path = os.path.join(outdir, "PlayerBaseStat.json")

    export_json(parsed_rows[0], PLAYER_STAT_SCHEMA, json_path)

    print("변환 완료 (기본 스탯 1건)")
    print(f"  - JSON: {json_path}")
    return 0


# ---------------------------------------------------------------------------
# 9. CLI
# ---------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="플레이어 기본 스탯 xlsx 템플릿 생성 / json 변환 도구")
    subparsers = parser.add_subparsers(dest="command", required=True)

    p_template = subparsers.add_parser("template", help="xlsx 템플릿 생성")
    p_template.add_argument("--output", default="PlayerBaseStat.xlsx", help="생성할 xlsx 경로")

    p_convert = subparsers.add_parser("convert", help="xlsx -> json 변환")
    p_convert.add_argument("--input", required=True, help="입력 xlsx 파일 경로")
    p_convert.add_argument("--outdir", default="./output", help="출력 디렉토리 (기본: ./output)")

    args = parser.parse_args()

    if args.command == "template":
        build_template(args.output)
        sys.exit(0)
    elif args.command == "convert":
        sys.exit(run_convert(args.input, args.outdir))


if __name__ == "__main__":
    main()
